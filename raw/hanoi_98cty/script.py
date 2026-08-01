import glob
from pathlib import Path

import openpyxl
import pandas as pd

BASE_DIR = Path(__file__).resolve().parent
# ponytail: dùng glob thay vì gõ thẳng tên file tiếng Việt trong source — filename trên đĩa
# dùng Unicode NFD (dấu tổ hợp), còn literal gõ tay/editor thường lưu NFC -> lệch byte, FileNotFoundError.
EXCEL_PATH = Path(glob.glob(str(BASE_DIR / "*.xlsx"))[0])
INVOICE_OUTPUT = BASE_DIR / "invoice.csv"
COMPANY_PATH = BASE_DIR / "company.csv"
mst_chinh = "0109082787"  # AZURA — công ty có nhãn AI thật (HĐ_Đầu_Vào / HĐ_Bán_Ra)

META_SHEETS = {"100MST", "2021", "BCTC_2021", "HĐ_Đầu_Vào", "HĐ_Bán_Ra"}

COMPANY_COLS = ["mst", "ten_cong_ty", "linh_vuc", "dia_chi", "doanh_thu", "nam_bao_cao"]
INVOICE_COLS = ["so_hoa_don", "ngay_xuat", "mst_nguon", "mst_dich", "mo_ta",
                "tien_chua_thue", "thue_gtgt", "loai_gd", "nhan_ai"]


def normalize_mst(value: object) -> str:
    if pd.isna(value):
        return ""
    digits = "".join(ch for ch in str(value).strip() if ch.isdigit())
    return digits.zfill(10)


def normalize_money(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0).astype("int64")


def normalize_date(value: object) -> str:
    if pd.isna(value):
        return ""

    text = str(value).strip()
    numeric_text = text.replace(".0", "")
    if numeric_text.isdigit():
        serial = int(numeric_text)
        if 20000 <= serial <= 60000:
            return (pd.Timestamp("1899-12-30") + pd.to_timedelta(serial, unit="D")).strftime("%Y-%m-%d")
        return text

    if len(text) == 10 and text[4] == "-" and text[7] == "-":
        parsed = pd.to_datetime(text, format="%Y-%m-%d", errors="coerce")
        if pd.notna(parsed):
            return parsed.strftime("%Y-%m-%d")

    if len(text) == 10 and text[2] == "/" and text[5] == "/":
        parsed = pd.to_datetime(text, format="%d/%m/%Y", errors="coerce")
        if pd.notna(parsed):
            return parsed.strftime("%Y-%m-%d")

    if " " in text or "-" in text or "/" in text:
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=("/" in text))
        if pd.notna(parsed):
            return parsed.strftime("%Y-%m-%d")

    return text


def is_valid_mst(mst: str) -> bool:
    return bool(mst) and len(mst) == 10 and mst.isdigit() and mst != "0000000000"


# =============================================
# PHẦN 1 — HĐ_Đầu_Vào / HĐ_Bán_Ra: nguồn nhãn AI thật (chỉ xoay quanh AZURA)
# =============================================
df_mua = pd.read_excel(EXCEL_PATH, sheet_name="HĐ_Đầu_Vào", header=3)
df_mua_clean = df_mua.rename(columns={
    "Số hóa đơn": "so_hoa_don", "Ngày xuất HĐ": "ngay_xuat",
    "MST người bán": "mst_nguon", "Tên NCC": "ten_nguon",
    "Nội dung hóa đơn": "mo_ta", "Tiền chưa thuế (VNĐ)": "tien_chua_thue",
    "Tiền thuế GTGT": "thue_gtgt", "Nhãn AI (label)": "nhan_ai",
}).copy()
df_mua_clean["mst_nguon"] = df_mua_clean["mst_nguon"].apply(normalize_mst)
df_mua_clean["mst_dich"] = mst_chinh
df_mua_clean["loai_gd"] = "mua_vao"

df_ban = pd.read_excel(EXCEL_PATH, sheet_name="HĐ_Bán_Ra", header=4)
df_ban_clean = df_ban.rename(columns={
    "Số hóa đơn": "so_hoa_don", "Ngày xuất HĐ": "ngay_xuat",
    "MST Khách hàng": "mst_dich", "Tên khách hàng": "ten_dich",
    "Nội dung hàng hóa / dịch vụ": "mo_ta", "Tiền chưa thuế (VNĐ)": "tien_chua_thue",
    "Tiền thuế GTGT": "thue_gtgt", "Nhãn AI": "nhan_ai",
}).copy()
df_ban_clean["mst_nguon"] = mst_chinh
df_ban_clean["mst_dich"] = df_ban_clean["mst_dich"].apply(normalize_mst)
df_ban_clean["loai_gd"] = "ban_ra"

labeled_df = pd.concat([df_mua_clean[INVOICE_COLS], df_ban_clean[INVOICE_COLS]], ignore_index=True)
labeled_df["mst_nguon"] = labeled_df["mst_nguon"].apply(normalize_mst)
labeled_df["mst_dich"] = labeled_df["mst_dich"].apply(normalize_mst)
labeled_df["ngay_xuat"] = labeled_df["ngay_xuat"].apply(normalize_date)
labeled_df["nhan_ai"] = pd.to_numeric(labeled_df["nhan_ai"], errors="coerce").fillna(0).astype("int64")
labeled_df = labeled_df[labeled_df["so_hoa_don"].notna()]
labeled_df = labeled_df[~labeled_df["so_hoa_don"].astype(str).str.contains("^STT$", case=False, na=False)]
print(f"Số hoá đơn có nhãn AI thật (HĐ_Đầu_Vào/HĐ_Bán_Ra, AZURA): {len(labeled_df)}")

# NOTE: AZURA (mst_chinh) KHÔNG có sheet riêng trong 86 sheet công ty — toàn bộ hoá đơn của nó
# chỉ có ở đây. Ban đầu có ý định đối chiếu (mst_nguon, mst_dich, ngay_xuat) để "chuyển" nhãn AI
# sang các dòng trích từ sheet công ty, nhưng 2 nguồn này là 2 lớp dữ liệu độc lập (khác cách đánh
# số hoá đơn, khác ngày) nên gần như không khớp được — nên bỏ hướng đó, UNION thẳng labeled_df vào
# làm cạnh riêng (nó đã tự mang đúng nhãn nhan_ai của nó rồi, không cần đối chiếu).

# =============================================
# PHẦN 2 — Duyệt 86 sheet công ty: dựng đầy đủ mạng lưới hoá đơn
# =============================================
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
company_sheets = [name for name in wb.sheetnames if name not in META_SHEETS]
print(f"Số sheet công ty tìm thấy: {len(company_sheets)}")

edges = []
name_hints = {}  # mst đối tác -> tên (để đặt tên cho công ty ngoài danh sách 86)

for sheet_name in company_sheets:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    sheet_mst = None
    for row in rows[:6]:
        if row[0] == "MST:":
            sheet_mst = normalize_mst(row[1])
            break
    if not sheet_mst:
        sheet_mst = normalize_mst(sheet_name)  # fallback: tên sheet chính là MST

    def extract_section(marker: str, mst_col_is_source: bool):
        start = None
        for i, row in enumerate(rows):
            if row[0] == marker:
                start = i + 1  # dòng header nằm ngay sau marker
                break
        if start is None:
            return
        for row in rows[start + 1:]:
            if row[0] is None:
                continue
            if isinstance(row[0], str) and row[0].strip().lower() == "tổng cộng":
                break
            counterparty_mst = normalize_mst(row[3])
            if not counterparty_mst:
                continue
            counterparty_name = row[4]
            if counterparty_name:
                name_hints.setdefault(counterparty_mst, counterparty_name)

            mst_nguon = counterparty_mst if mst_col_is_source else sheet_mst
            mst_dich = sheet_mst if mst_col_is_source else counterparty_mst
            edges.append({
                "so_hoa_don": str(row[1]),
                "ngay_xuat": normalize_date(row[2]),
                "mst_nguon": mst_nguon,
                "mst_dich": mst_dich,
                "mo_ta": row[5],
                "tien_chua_thue": row[6],
                "thue_gtgt": row[8],
                "loai_gd": "mua_vao" if mst_col_is_source else "ban_ra",
                "nhan_ai": 0,  # sheet công ty thường không có nhãn AI thật, xem PHẦN 1 cho AZURA
            })

    extract_section("Hóa đơn đầu vào", mst_col_is_source=True)   # counterparty = mst_nguon (seller)
    extract_section("Hóa đơn đầu ra", mst_col_is_source=False)   # counterparty = mst_dich (buyer)

edge_df = pd.DataFrame(edges, columns=INVOICE_COLS)
print(f"Số cạnh trích từ 86 sheet công ty (trước gộp + khử trùng lặp): {len(edge_df)}")

# Gộp thêm các hoá đơn có nhãn AI thật của AZURA (nguồn riêng, xem PHẦN 1)
edge_df = pd.concat([edge_df, labeled_df], ignore_index=True)

edge_df["mst_nguon"] = edge_df["mst_nguon"].apply(normalize_mst)
edge_df["mst_dich"] = edge_df["mst_dich"].apply(normalize_mst)
edge_df["tien_chua_thue"] = normalize_money(edge_df["tien_chua_thue"])
edge_df["thue_gtgt"] = normalize_money(edge_df["thue_gtgt"])
edge_df["nhan_ai"] = pd.to_numeric(edge_df["nhan_ai"], errors="coerce").fillna(0).astype("int64")

edge_df = edge_df[edge_df["mst_nguon"].apply(is_valid_mst) & edge_df["mst_dich"].apply(is_valid_mst)]

# Khử trùng lặp 2 chiều: cùng 1 hoá đơn được ghi ở CẢ sheet bên bán ("đầu ra") LẪN sheet
# bên mua ("đầu vào") khi cả 2 công ty đều nằm trong danh sách 86 — chỉ giữ 1 bản ghi.
before_dedupe = len(edge_df)
edge_df = edge_df.drop_duplicates(subset=["so_hoa_don", "mst_nguon", "mst_dich"], keep="first")
print(f"Đã khử {before_dedupe - len(edge_df)} dòng trùng lặp 2 chiều")
print(f"Tổng số cạnh hoá đơn cuối cùng: {len(edge_df)}")

# NebulaGraph định danh 1 cạnh bằng (src, edge_type, dst, rank) — rank mặc định = 0 nếu không set.
# Một cặp công ty (A, B) có thể có NHIỀU hoá đơn trong năm -> nếu không có rank, các hoá đơn sau
# sẽ ghi đè hoá đơn trước (cùng src/dst/rank=0) và mất dữ liệu. Rank ở đây chỉ là số thứ tự tăng
# dần, không mang ý nghĩa nghiệp vụ — chỉ để đảm bảo mỗi hoá đơn là 1 cạnh riêng biệt.
edge_df = edge_df.reset_index(drop=True)
edge_df["rank"] = edge_df.index

edge_df.to_csv(INVOICE_OUTPUT, index=False, header=False, encoding="utf-8")
print(f"Xuất xong {len(edge_df)} hoá đơn -> {INVOICE_OUTPUT.name}")

# =============================================
# PHẦN 3 — Mở rộng company.csv: chuẩn hoá 86 công ty gốc + thêm công ty ngoài danh sách
# =============================================
company_df = pd.read_csv(COMPANY_PATH, header=None, names=COMPANY_COLS, encoding="utf-8-sig", dtype=str)
company_df["mst"] = company_df["mst"].apply(normalize_mst)
company_df["doanh_thu"] = pd.to_numeric(company_df["doanh_thu"], errors="coerce").fillna(0).astype("int64")

known_mst = set(company_df["mst"])
all_edge_mst = set(edge_df["mst_nguon"]) | set(edge_df["mst_dich"])
external_mst = sorted(all_edge_mst - known_mst)

print(f"Số MST xuất hiện trong hoá đơn nhưng KHÔNG nằm trong 86 công ty: {len(external_mst)}")
if external_mst:
    ext_rows = pd.DataFrame({
        "mst": external_mst,
        "ten_cong_ty": [name_hints.get(m, "(Chưa rõ tên)") for m in external_mst],
        "linh_vuc": [""] * len(external_mst),
        "dia_chi": [""] * len(external_mst),
        "doanh_thu": [0] * len(external_mst),
        "nam_bao_cao": [""] * len(external_mst),
    })
    company_df = pd.concat([company_df, ext_rows], ignore_index=True)

company_df.to_csv(COMPANY_PATH, index=False, header=False, encoding="utf-8")
print(f"Tổng công ty sau mở rộng: {len(company_df)} -> {COMPANY_PATH.name}")
